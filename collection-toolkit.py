#!/usr/bin/env python
# Author: Daniel Teycheney

# Import Modules
from nornir import InitNornir
from nornir.plugins.tasks.networking import napalm_get
import requests
import pathlib
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import datetime as dt
import openpyxl

# Disable urllib3 warnings
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


"""
The following five functions are used to retrieve NAPALM getters required
for the summary spreadsheet
"""


def get_facts(task):
    task.run(name="Get facts", task=napalm_get, getters=["facts"])
    return "Complete"


def get_interfaces(task):
    task.run(name="Get interfaces", task=napalm_get, getters=["interfaces"])
    return "Complete"


def get_interfaces_ip(task):
    task.run(name="Get interfaces IP", task=napalm_get, getters=["interfaces_ip"])
    return "Complete"


def get_lldp_neighbors(task):
    task.run(name="Get LLDP neighbors", task=napalm_get, getters=["lldp_neighbors"])
    return "Complete"


def get_users(task):
    task.run(name="Get users", task=napalm_get, getters=["users"])
    return "Complete"


def main_collector(wb, log_file):
    """
    This is the main function of the application. In this function, we run tasks against all hosts
    in the inventory and parse the results and place them into various spreadsheet tabs.

    There are five spreadsheets and getter that we are collecting

    Facts - The facts about the hosts
    Interfaces - A list of interfaces on each host
    Interfaces_IP - A list of interfaces with IP addressed on each host
    LLDP - A list of LLDP neighbors on each host
    Users - A list of local usernames on each host
    :param wb: The Excel workbook where the results will be saved to.
    :param log_file: The log file which will save the results as we process through the host.
    :return:
    """
    """
    The following block of code creates the various spreadsheet tabs, assign headers to those
    spreadsheets and inserts those headers at the top of that spreadsheet.
    """
    # Create facts worksheet
    facts_ws = wb.create_sheet("Facts")
    # Statically assign headers
    facts_headers = [
        "Hostname",
        "Vendor",
        "Model",
        "OS Version",
        "Serial Number",
        "Uptime",
    ]
    # Write headers on the top line of the file
    facts_ws.append(facts_headers)
    # Create interfaces worksheet
    interfaces_ws = wb.create_sheet("Interfaces")
    # Statically assign headers
    interfaces_headers = [
        "Name",
        "Interface Name",
        "Interface Description",
        "Interface Up",
        "Interface Enabled",
    ]
    # Write headers on the top line of the file
    interfaces_ws.append(interfaces_headers)
    # Create interfaces IP worksheet
    interfaces_ip_ws = wb.create_sheet("Interfaces_IP")
    # Statically assign headers
    interfaces_ip_headers = [
        "Name",
        "Interface Name",
        "IPv4 Address",
        "IPv4 Prefix Length",
        "IPv6 Address",
        "IPv6 Prefix Length",
    ]
    # Write headers on the top line of the file
    interfaces_ip_ws.append(interfaces_ip_headers)
    # Create LLDP neighbors worksheet
    lldp_nei_ws = wb.create_sheet("LLDP")
    # Statically assign headers
    lldp_nei_headers = [
        "Local Hostname",
        "Local Port",
        "Remote Hostname",
        "Remote Port",
    ]
    # Write headers on the top line of the file
    lldp_nei_ws.append(lldp_nei_headers)
    # Create Users worksheet
    users_ws = wb.create_sheet("Users")
    # Statically assign headers
    users_headers = ["Hostname", "Username", "Level", "Password", "SSH Keys"]
    # Write headers on the top line of the file
    users_ws.append(users_headers)
    # Initialize Nornir and define the inventory variables.
    nr = InitNornir(
        inventory={
            "options": {
                "host_file": "inventory/hosts.yaml",
                "group_file": "inventory/groups.yaml",
                "defaults_file": "inventory/defaults.yaml",
            }
        }
    )
    """
    The following block of code assigns a filter based on
    platform to a variable. This variable is used later on
    to apply logic in for loops
    """
    ios_devices = nr.filter(platform="ios")
    junos_devices = nr.filter(platform="junos")
    eos_devices = nr.filter(platform="eos")
    nxos_devices = nr.filter(platform="nxos")
    iosxr_devices = nr.filter(platform="iosxr")
    """
    Executing the get_interfaces task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_interfaces = ios_devices.run(name="Processing interfaces", task=get_interfaces)
    junos_interfaces = junos_devices.run(
        name="Processing interfaces", task=get_interfaces
    )
    eos_interfaces = eos_devices.run(name="Processing interfaces", task=get_interfaces)
    nxos_interfaces = nxos_devices.run(
        name="Processing interfaces", task=get_interfaces
    )
    iosxr_interfaces = iosxr_devices.run(
        name="Processing interfaces", task=get_interfaces
    )
    # Take all the those results and add them to a list so we can iterate over the result
    os_interfaces = [
        ios_interfaces,
        junos_interfaces,
        eos_interfaces,
        nxos_interfaces,
        iosxr_interfaces,
    ]
    # Iterate over the results in the list above
    for os in os_interfaces:
        # For loop to process individual results
        for host, task_results in os.items():
            # Display printout
            print("Start Processing Host - Interfaces: " + str(host) + "\n")
            # Add to log file
            log_file.write("Start Processing Host - Interfaces: " + str(host) + "\n")
            # Extract the result of the task
            get_interfaces_result = task_results[1].result
            interface_name_result = get_interfaces_result["interfaces"]
            # Empty list which will be appended to in for loop
            int_list = []
            # For loop to retrieve the list of interfaces
            for entry in interface_name_result:
                # Append entries to the int_list list
                int_list.append(entry)
            # For loop to loop through list of interfaces and extract interface values
            for int in int_list:
                # Assign individual interface entry to a variable
                int_result = interface_name_result[int]
                # Extract the interface description and assign to a variable
                int_desc_result = int_result["description"]
                # Extract the interface state and assign to a variable
                int_up_result = int_result["is_up"]
                # Extract the whether the interface is enabled and assign to a variable
                int_enable_result = int_result["is_enabled"]
                # Display printout
                print("Interface Name: " + str(int))
                # Add to log file
                log_file.write("Interface Name: " + str(int) + "\n")
                # Display printout
                print("Interface Description: " + str(int_desc_result))
                # Add to log file
                log_file.write("Interface Description: " + str(int_desc_result) + "\n")
                # Display printout
                print("Interface Up: " + str(int_up_result))
                # Add to log file
                log_file.write("Interface Up: " + str(int_up_result) + "\n")
                # Display printout
                print("Interface Enabled: " + str(int_enable_result))
                # Add to log file
                log_file.write("Interface Enabled: " + str(int_enable_result) + "\n")
                line = [host, int, int_desc_result, int_up_result, int_enable_result]
                # Write values to file
                interfaces_ws.append(line)
            # Display printout
            print("End Processing Host - Interfaces: " + str(host) + "\n")
            # Add to log file
            log_file.write("End Processing Host - Interfaces: " + str(host) + "\n\n")
    """
    Executing the get_facts task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_facts = ios_devices.run(name="Processing facts", task=get_facts)
    junos_facts = junos_devices.run(name="Processing facts", task=get_facts)
    eos_facts = eos_devices.run(name="Processing facts", task=get_facts)
    nxos_facts = nxos_devices.run(name="Processing facts", task=get_facts)
    iosxr_facts = iosxr_devices.run(name="Processing facts", task=get_facts)
    # Take all the those results and add them to a list so we can iterate over the result
    os_facts = [ios_facts, junos_facts, eos_facts, nxos_facts, iosxr_facts]
    # Iterate over the results in the list above
    for os in os_facts:
        # For loop to process individual results
        for host, task_results in os.items():
            # Display printout
            print("Start Processing Host - Facts: " + str(host) + "\n")
            # Add to log file
            log_file.write("Start Processing Host - Facts: " + str(host) + "\n")
            # Extract the result of the task
            get_facts_result = task_results[1].result
            # Extract the Vendor and assign to a variable
            vendor_result = get_facts_result["facts"]["vendor"]
            # Extract the Model and assign to a variable
            model_result = get_facts_result["facts"]["model"]
            # Extract the OS Version and assign to a variable
            version_result = get_facts_result["facts"]["os_version"]
            # Extract the Serial Number and assign to a variable
            ser_num_result = get_facts_result["facts"]["serial_number"]
            # Extract the Uptime and assign to a variable
            uptime_result = get_facts_result["facts"]["uptime"]
            # Display printout
            print("Vendor: " + str(vendor_result))
            # Add to log file
            log_file.write("Vendor: " + str(vendor_result) + "\n")
            # Display printout
            print("Model: " + str(model_result))
            # Add to log file
            log_file.write("Model: " + str(model_result) + "\n")
            # Display printout
            print("OS Version: " + str(version_result))
            # Add to log file
            log_file.write("OS Version: " + str(version_result) + "\n")
            # Display printout
            print("Serial Number: " + str(version_result))
            # Add to log file
            log_file.write("Serial Number: " + str(version_result) + "\n")
            # Display printout
            print("Uptime: " + str(version_result))
            # Add to log file
            log_file.write("Uptime: " + str(version_result) + "\n")
            line = [
                host,
                vendor_result,
                model_result,
                version_result,
                ser_num_result,
                uptime_result,
            ]
            # Debug print
            # print(line)
            # Write values to file
            facts_ws.append(line)
            # Display printout
            print("End Processing Host - Facts: " + str(host) + "\n")
            # Add to log file
            log_file.write("End Processing Host - Facts: " + str(host) + "\n\n")
    """
    Executing the get_interfaces_ip task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_interfaces_ip = ios_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    junos_interfaces_ip = junos_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    eos_interfaces_ip = eos_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    nxos_interfaces_ip = nxos_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    iosxr_interfaces_ip = iosxr_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    # Take all the those results and add them to a list so we can iterate over the result
    os_interfaces_ip = [
        ios_interfaces_ip,
        junos_interfaces_ip,
        eos_interfaces_ip,
        nxos_interfaces_ip,
        iosxr_interfaces_ip,
    ]
    # Iterate over the results in the list above
    for os in os_interfaces_ip:
        # For loop to process individual results
        for host, task_results in os.items():
            # Display printout
            print("Start Processing Host - Interfaces IP: " + str(host) + "\n")
            # Add to log file
            log_file.write("Start Processing Host - Interfaces IP: " + str(host) + "\n")
            # Gather results from task
            get_interfaces_ip_result = task_results[1].result
            # Filter the results
            interface_ip_name_result = get_interfaces_ip_result["interfaces_ip"]
            # Empty list which will be appended to in for loop
            int_ip_list = []
            # For loop to retrieve the list of interfaces
            for entry in interface_ip_name_result:
                # Append entries to the int_list list
                int_ip_list.append(entry)
                # Debug print
                # print(int_ip_list)
            # For loop to loop through list of IPv4 interfaces and extract interface_ip values
            for int_ip in int_ip_list:
                # Assign individual interface entry to a variable
                final_int_ip = interface_ip_name_result[int_ip]
                # Assign IPv4 address to a variable
                int_ipv4_addr = final_int_ip["ipv4"]
                # Debug print
                # print(int_ip)
                # For loop to extract single IPv4 address
                for ip in int_ipv4_addr.items():
                    # Assign IPv4 address to a variable
                    ipv4_address = ip[0]
                    # Debug print
                    # print(ipv4_address)
                    # For loop to extract prefix length from prefix_length variable
                    for key, prefix_length_v4 in ip[1].items():
                        # Print must be left on or for loop isn't activated.
                        print("Prefix length debug print - Ignore")
                        # Debug print
                        # print(prefix_length)
                # Try/Except block to look handle IPv6 addresses, namely when they are not there.
                try:
                    # Assign IPv6 address to a variable
                    int_ipv6_addr = final_int_ip["ipv6"]
                    for ip in int_ipv6_addr.items():
                        # Assign IPv6 address to a variable
                        ipv6_address = ip[0]
                        # Debug print
                        # print(ipv6_address)
                        # For loop to extract prefix length from prefix_length variable
                        for key, prefix_length_v6 in ip[1].items():
                            # Print must be left on or for loop isn't activated.
                            print("Prefix length debug print - Ignore")
                            # Debug print
                            # print(prefix_length)
                # When the IPv6 address is not there, it throws a key error
                except KeyError:
                    # Display printout
                    print("IPv6 Address not configured")
                    # Add to log file
                    log_file.write("IPv6 Address not configured" + "\n")
                    # Override value so there is a result which is clear that it is not configured.
                    ipv6_address = "NOT CONFIGURED"
                    # Override value so there is a result which is clear that it is not configured.
                    prefix_length_v6 = "NOT CONFIGURED"
                # Display printout
                print("Interface Name: " + str(int_ip))
                # Add to log file
                log_file.write("Interface Name: " + str(int_ip) + "\n")
                # Display printout
                print("IPv4 Address: " + str(ipv4_address))
                # Add to log file
                log_file.write("IPv4 Address: " + str(ipv4_address) + "\n")
                # Display printout
                print("IPv4 Prefix Length: " + str(prefix_length_v4))
                # Add to log file
                log_file.write("IPv4 Prefix Length: " + str(prefix_length_v4) + "\n")
                # Display printout
                print("IPv6 Address: " + str(ipv6_address))
                # Add to log file
                log_file.write("IPv6 Address: " + str(ipv6_address) + "\n")
                # Display printout
                print("IPv6 Prefix Length: " + str(prefix_length_v6))
                # Add to log file
                log_file.write("IPv6 Prefix Length: " + str(prefix_length_v6) + "\n")
                # Append results to a line to be saved to the workbook
                line = [
                    host,
                    int_ip,
                    str(ipv4_address),
                    str(prefix_length_v4),
                    str(ipv6_address),
                    str(prefix_length_v6),
                ]
                # Save values to row in workbook
                interfaces_ip_ws.append(line)
            # Display printout
            print("End Processing Host - Interfaces IP: " + str(host) + "\n")
            # Add to log file
            log_file.write("End Processing Host - Interfaces IP: " + str(host) + "\n\n")
    """
    Executing the get_lldp_neighbors task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_lldp = ios_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    junos_lldp = junos_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    eos_lldp = eos_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    nxos_lldp = nxos_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    iosxr_lldp = iosxr_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    # Take all the those results and add them to a list so we can iterate over the result
    os_lldp = [ios_lldp, junos_lldp, eos_lldp, nxos_lldp, iosxr_lldp]
    # Iterate over the results in the list above
    for os in os_lldp:
        # For loop to process individual results
        for host, task_results in os.items():
            # Display printout
            print("Start Processing Host - LLDP: " + str(host) + "\n")
            # Add to log file
            log_file.write("Start Processing Host - LLDP: " + str(host) + "\n")
            # Extract the result of the task
            lldp_nei_result = task_results[1].result
            lldp_nei_name_result = lldp_nei_result["lldp_neighbors"]
            # Empty list which will be appended to in for loop
            neighbor_list = []
            # For loop to retrieve the list of interfaces
            for entry in lldp_nei_name_result:
                # Append entries to the neighbor_list list
                neighbor_list.append(entry)
                # Debug print
                # print(neighbor_list)
            for local_port in neighbor_list:
                # Extract the remote port and assign to a variable
                remote_port = lldp_nei_name_result[local_port][0]["port"]
                # Extract the remote username and assign to a variable
                remote_hostname = lldp_nei_name_result[local_port][0]["hostname"]
                # Display printout
                print("Local Port: " + str(local_port))
                # Add to log file
                log_file.write("Local Port: " + str(local_port) + "\n")
                # Display printout
                print("Remote Port: " + str(remote_port))
                # Add to log file
                log_file.write("Remote Port: " + str(remote_port) + "\n")
                # Display printout
                print("Remote Hostname: " + str(remote_hostname))
                # Add to log file
                log_file.write("Remote Hostname: " + str(remote_hostname) + "\n")
                # Append results to a line to be saved to the workbook
                line = [host, local_port, remote_hostname, remote_port]
                # Write values to file
                lldp_nei_ws.append(line)
            # Display printout
            print("End Processing Host - LLDP: " + str(host) + "\n")
            # Add to log file
            log_file.write("End Processing Host - LLDP: " + str(host) + "\n\n")
    """
    Executing the get_interfaces task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_users = ios_devices.run(name="Processing users", task=get_users)
    junos_users = junos_devices.run(name="Processing users", task=get_users)
    eos_users = eos_devices.run(name="Processing users", task=get_users)
    nxos_users = nxos_devices.run(name="Processing users", task=get_users)
    iosxr_users = iosxr_devices.run(name="Processing users", task=get_users)
    # Take all the those results and add them to a list so we can iterate over the result
    os_users = [
        ios_users,
        # TODO: Need to work out this junos_users filter not working.
        # junos_users,
        eos_users,
        nxos_users,
        iosxr_users,
    ]
    # Iterate over the results in the list above
    for os in os_users:
        # For loop to process individual results
        for host, task_results in os.items():
            # Display printout
            print("Start Processing Host - Users: " + str(host) + "\n")
            # Add to log file
            log_file.write("Start Processing Host - Users: " + str(host) + "\n")
            # Extract the result of the task
            get_users_result = task_results[1].result
            users_name_result = get_users_result["users"]
            # print(users_name_result)
            # Empty list which will be appended to in for loop
            user_list = []
            for entry in users_name_result:
                # Append entries to the neighbor_list list
                user_list.append(entry)
                # print(user_list)
            for user in user_list:
                # Extract the User privilege level and assign to a variable
                user_level = users_name_result[user]["level"]
                # Extract the User password and assign to a variable
                user_pw = users_name_result[user]["password"]
                # Extract the SSH keys and assign to a variable
                user_ssh = users_name_result[user]["sshkeys"]
                # Display printout
                print("Username: " + str(user))
                # Add to log file
                log_file.write("Username: " + str(user) + "\n")
                # Display printout
                print("Level: " + str(user_level))
                # Add to log file
                log_file.write("Level: " + str(user_level) + "\n")
                # Display printout
                print("Password: " + str(user_pw))
                # Add to log file
                log_file.write("Password: " + str(user_pw) + "\n")
                # Display printout
                print("SSH Keys: " + str(user_ssh))
                # Add to log file
                log_file.write("SSH Keys: " + str(user_ssh) + "\n")
                # Append results to a line to be saved to the workbook
                line = [host, user, user_level, user_pw, str(user_ssh)]
                # # Write values to file
                users_ws.append(line)
            # Display printout
            print("End Processing Host - Users: " + str(host) + "\n")
            # Add to log file
            log_file.write("End Processing Host - Users: " + str(host) + "\n")
    # JUNOS Platform Block
    """
    I am not sure how this is working given that the task results
    are failing, but it is....
    """
    for host, task_results in junos_users.items():
        # Display printout
        print("Start Processing Host - Users: " + str(host) + "\n")
        # Add to log file
        log_file.write("Start Processing Host - Users: " + str(host) + "\n")
        get_users_result = task_results.result
        # print(get_users_result)
        # users_name_result = get_users_result["users"]
        # # print(users_name_result)
        # Empty list which will be appended to in for loop
        user_list = []
        for entry in users_name_result:
            # Append entries to the neighbor_list list
            user_list.append(entry)
            # print(user_list)
        for user in user_list:
            # Extract the User privilege level and assign to a variable
            user_level = users_name_result[user]["level"]
            # Extract the User password and assign to a variable
            user_pw = users_name_result[user]["password"]
            # Extract the SSH keys and assign to a variable
            user_ssh = users_name_result[user]["sshkeys"]
            # Display printout
            print("Username: " + str(user))
            # Add to log file
            log_file.write("Username: " + str(user) + "\n")
            # Display printout
            print("Level: " + str(user_level))
            # Add to log file
            log_file.write("Level: " + str(user_level) + "\n")
            # Display printout
            print("Password: " + str(user_pw))
            # Add to log file
            log_file.write("Password: " + str(user_pw) + "\n")
            # Display printout
            print("SSH Keys: " + str(user_ssh))
            # Add to log file
            log_file.write("SSH Keys: " + str(user_ssh) + "\n")
            # Append results to a line to be saved to the workbook
            line = [host, user, user_level, user_pw, str(user_ssh)]
            # # Write values to file
            users_ws.append(line)
        # Display printout
        print("End Processing Host - Users: " + str(host) + "\n")
        # Add to log file
        log_file.write("End Processing Host - Users: " + str(host) + "\n")


def process_functions(wb, log_file):
    main_collector(wb, log_file)


def create_workbook():
    # Capture time
    cur_time = dt.datetime.now()
    # Cleanup time, so that the format is clean for the output file 2016-12-01-13-04-59
    fmt_time = cur_time.strftime("%Y-%m-%d-%H-%M-%S")
    # Set log directory variable
    log_dir = "logs"
    # Create log directory if it doesn't exist.
    pathlib.Path(log_dir).mkdir(exist_ok=True)
    # Create log file name, with timestamp in the name
    filename = str("COLLECTION-LOG") + "-" + fmt_time + ".txt"
    # Join the log file name and log directory together into a variable
    log_file_path = log_dir + "/" + filename
    # Create the log file
    log_file = open(log_file_path, "w")
    # Setup workbook parameters
    wb = openpyxl.Workbook()
    # Execute program
    process_functions(wb, log_file)
    customer_name = "Customer"
    time_now = (dt.datetime.now()).strftime("%d-%m-%H-%M")
    wb_name = "Diagnostics-" + customer_name + "-2019-" + time_now + ".xlsx"
    print(wb_name)
    wb.save(wb_name)


# Execute main function
create_workbook()
